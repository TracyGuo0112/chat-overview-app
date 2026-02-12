#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
为Excel文件的每个sheet添加翻译列
将remark列的内容翻译成中文
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import time
import re
from googletrans import Translator
import warnings
warnings.filterwarnings('ignore')

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class RemarkTranslator:
    def __init__(self):
        self.translator = Translator()
        self.translation_cache = {}
        
    def detect_language(self, text):
        """检测文本语言"""
        if not text or pd.isna(text) or str(text).strip() == '':
            return 'unknown'
        
        text_str = str(text).strip()
        
        # 简单的语言检测规则
        # 检测中文
        if re.search(r'[\u4e00-\u9fff]', text_str):
            return 'zh'
        
        # 检测韩文
        if re.search(r'[\uac00-\ud7af]', text_str):
            return 'ko'
        
        # 检测日文
        if re.search(r'[\u3040-\u309f\u30a0-\u30ff]', text_str):
            return 'ja'
        
        # 其他假设为英文或欧洲语言
        return 'auto'
    
    def translate_text(self, text, max_retries=3):
        """翻译文本到中文"""
        if not text or pd.isna(text) or str(text).strip() == '':
            return ''
        
        text_str = str(text).strip()
        
        # 检查缓存
        if text_str in self.translation_cache:
            return self.translation_cache[text_str]
        
        # 检测语言
        lang = self.detect_language(text_str)
        
        # 如果已经是中文，直接返回
        if lang == 'zh':
            self.translation_cache[text_str] = text_str
            return text_str
        
        # 翻译
        for attempt in range(max_retries):
            try:
                # 限制文本长度，避免翻译API错误
                if len(text_str) > 500:
                    text_str = text_str[:500] + '...'
                
                result = self.translator.translate(text_str, dest='zh')
                translated = result.text
                
                # 缓存结果
                self.translation_cache[text_str] = translated
                
                # 添加延迟避免API限制
                time.sleep(0.1)
                
                return translated
                
            except Exception as e:
                logger.warning(f"翻译失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(1)  # 重试前等待
                else:
                    # 最后一次尝试失败，返回原文
                    logger.error(f"翻译最终失败，返回原文: {text_str[:50]}...")
                    return f"[翻译失败] {text_str}"
        
        return text_str

def add_translation_to_excel():
    """为Excel文件的每个sheet添加翻译列"""
    
    input_file = '意图分类数据_按类别分组.xlsx'
    output_file = '意图分类数据_按类别分组_含翻译.xlsx'
    
    logger.info(f"开始处理文件: {input_file}")
    
    # 创建翻译器
    translator = RemarkTranslator()
    
    # 读取原Excel文件
    workbook = openpyxl.load_workbook(input_file)
    sheet_names = workbook.sheetnames
    
    logger.info(f"发现 {len(sheet_names)} 个sheet")
    
    total_translated = 0
    
    for sheet_name in sheet_names:
        logger.info(f"处理sheet: {sheet_name}")
        
        try:
            # 读取sheet数据
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            
            if 'remark' not in df.columns:
                logger.warning(f"Sheet '{sheet_name}' 没有remark列，跳过")
                continue
            
            # 添加翻译列
            translations = []
            
            for idx, row in df.iterrows():
                remark = row['remark']
                translation = translator.translate_text(remark)
                translations.append(translation)
                
                if (idx + 1) % 10 == 0:
                    logger.info(f"  已翻译 {idx + 1}/{len(df)} 条数据...")
            
            # 添加翻译列到DataFrame
            df['翻译'] = translations
            total_translated += len([t for t in translations if t])
            
            logger.info(f"  ✅ {sheet_name}: {len(translations)}条数据翻译完成")
            
        except Exception as e:
            logger.error(f"处理sheet '{sheet_name}' 时出错: {e}")
            continue
    
    workbook.close()
    
    # 重新创建Excel文件，包含翻译列
    logger.info("重新创建包含翻译的Excel文件...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        for sheet_name in sheet_names:
            try:
                # 读取原始数据
                df = pd.read_excel(input_file, sheet_name=sheet_name)
                
                if 'remark' in df.columns:
                    # 添加翻译
                    translations = []
                    for _, row in df.iterrows():
                        translation = translator.translate_text(row['remark'])
                        translations.append(translation)
                    
                    df['翻译'] = translations
                
                # 写入新文件
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 格式化
                worksheet = writer.sheets[sheet_name]
                
                # 设置表头格式
                header_font = Font(bold=True, color='FFFFFF')
                if sheet_name == '总览':
                    header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
                else:
                    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in worksheet[1]:  # 第一行是表头
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # 设置合适的列宽
                    if column_letter == worksheet.max_column:  # 翻译列
                        adjusted_width = min(max_length + 2, 60)  # 翻译列可以更宽
                    else:
                        adjusted_width = min(max_length + 2, 50)
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
            except Exception as e:
                logger.error(f"重新创建sheet '{sheet_name}' 时出错: {e}")
                continue
    
    logger.info(f"✅ 翻译完成！")
    logger.info(f"总计翻译: {total_translated} 条数据")
    logger.info(f"输出文件: {output_file}")
    
    return output_file

def main():
    """主函数"""
    try:
        logger.info("开始添加翻译列...")
        output_file = add_translation_to_excel()
        
        # 验证结果
        logger.info("验证翻译结果...")
        workbook = openpyxl.load_workbook(output_file)
        
        for sheet_name in workbook.sheetnames[:3]:  # 检查前3个sheet
            try:
                df = pd.read_excel(output_file, sheet_name=sheet_name)
                if '翻译' in df.columns and 'remark' in df.columns:
                    logger.info(f"✅ {sheet_name}: 包含翻译列")
                    # 显示样本
                    for i in range(min(2, len(df))):
                        if pd.notna(df.iloc[i]['remark']) and pd.notna(df.iloc[i]['翻译']):
                            remark = str(df.iloc[i]['remark'])[:50] + '...' if len(str(df.iloc[i]['remark'])) > 50 else str(df.iloc[i]['remark'])
                            translation = str(df.iloc[i]['翻译'])[:50] + '...' if len(str(df.iloc[i]['翻译'])) > 50 else str(df.iloc[i]['翻译'])
                            logger.info(f"  原文: {remark}")
                            logger.info(f"  译文: {translation}")
                            break
                else:
                    logger.warning(f"⚠️ {sheet_name}: 缺少翻译列或remark列")
            except Exception as e:
                logger.error(f"验证sheet '{sheet_name}' 时出错: {e}")
        
        workbook.close()
        
        print(f"\n🎉 翻译任务完成！")
        print(f"输出文件: {output_file}")
        print("每个sheet都已添加'翻译'列，包含remark的中文翻译")
        
    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}")
        raise

if __name__ == "__main__":
    main()
