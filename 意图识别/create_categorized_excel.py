#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据意图分类创建分类Excel文件
每个类别创建一个独立的sheet
"""

import pandas as pd
import json
import logging
from collections import defaultdict
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_categorized_excel():
    """创建按类别分类的Excel文件"""
    
    # 读取数据
    logger.info("读取标注数据...")
    df = pd.read_csv('用户需求意图识别_已标注.csv')
    
    # 读取排序后的类别信息
    with open('intent_categories_sorted.json', 'r', encoding='utf-8') as f:
        category_info = json.load(f)
    
    sorted_categories = category_info['sorted_categories']
    data_mapping = category_info['data_mapping']
    
    logger.info(f"总数据量: {len(df)}")
    logger.info(f"类别数量: {len(sorted_categories)}")
    
    # 创建Excel writer
    output_file = '意图分类数据_按类别分组.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 为每个类别创建sheet
        for i, (category, count) in enumerate(sorted_categories):
            sheet_name = f"{i}_{category}"
            logger.info(f"创建sheet: {sheet_name} ({count}条数据)")
            
            # 获取属于该类别的所有数据行
            category_rows = []
            
            for idx, row in df.iterrows():
                if pd.notna(row['意图分类']) and str(row['意图分类']).strip():
                    intents = [intent.strip() for intent in str(row['意图分类']).split(',')]
                    if category in intents:
                        category_rows.append(row)
            
            # 创建该类别的DataFrame
            if category_rows:
                category_df = pd.DataFrame(category_rows)
                # 重置索引
                category_df = category_df.reset_index(drop=True)
                
                # 写入Excel sheet
                category_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表以进行格式化
                worksheet = writer.sheets[sheet_name]
                
                # 设置表头格式
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in worksheet[1]:  # 第一行是表头
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # 设置合适的列宽，最大不超过50
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                logger.info(f"  ✅ {sheet_name}: {len(category_df)}条数据已写入")
            else:
                logger.warning(f"  ⚠️ {sheet_name}: 没有找到数据")
        
        # 创建总览sheet
        logger.info("创建总览sheet...")
        summary_data = []
        for i, (category, count) in enumerate(sorted_categories):
            percentage = count / len(df) * 100
            summary_data.append({
                '排序': i,
                '类别名称': category,
                '数据量': count,
                '占比(%)': f"{percentage:.1f}%",
                'Sheet名称': f"{i}_{category}"
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='总览', index=False)
        
        # 格式化总览sheet
        summary_worksheet = writer.sheets['总览']
        
        # 设置表头格式
        for cell in summary_worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 调整列宽
        column_widths = {'A': 8, 'B': 15, 'C': 10, 'D': 12, 'E': 20}
        for col, width in column_widths.items():
            summary_worksheet.column_dimensions[col].width = width
    
    logger.info(f"✅ Excel文件创建完成: {output_file}")
    
    # 验证结果
    logger.info("\n=== 创建结果验证 ===")
    workbook = openpyxl.load_workbook(output_file)
    sheet_names = workbook.sheetnames
    
    logger.info(f"总sheet数量: {len(sheet_names)}")
    logger.info("Sheet列表:")
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        row_count = sheet.max_row - 1  # 减去表头行
        logger.info(f"  {sheet_name}: {row_count}行数据")
    
    workbook.close()
    
    return output_file

if __name__ == "__main__":
    try:
        output_file = create_categorized_excel()
        print(f"\n🎉 Excel文件创建成功: {output_file}")
        print("文件包含以下内容:")
        print("- 总览sheet: 显示所有类别的统计信息")
        print("- 17个分类sheet: 每个类别的详细数据")
        print("- 按数据量降序排序的sheet命名")
        
    except Exception as e:
        print(f"❌ 创建Excel文件时出错: {e}")
        raise
