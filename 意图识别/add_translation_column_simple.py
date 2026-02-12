#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
为Excel文件的每个sheet添加翻译列
使用简单的翻译逻辑和常见短语翻译
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import re

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SimpleTranslator:
    def __init__(self):
        # 常见短语翻译字典
        self.translation_dict = {
            # 英文常见短语
            'terrible': '糟糕的',
            'bad': '不好的',
            'poor': '差的',
            'awful': '可怕的',
            'horrible': '糟糕的',
            'worst': '最差的',
            'good': '好的',
            'great': '很棒的',
            'excellent': '优秀的',
            'best': '最好的',
            'super': '超级的',
            'perfect': '完美的',
            'thank': '谢谢',
            'thanks': '谢谢',
            'hello': '你好',
            'hi': '嗨',
            'please': '请',
            'help': '帮助',
            'problem': '问题',
            'issue': '问题',
            'error': '错误',
            'broken': '坏了',
            'not work': '不工作',
            'doesnt work': '不工作',
            'repair': '维修',
            'fix': '修理',
            'buy': '购买',
            'purchase': '购买',
            'order': '订单',
            'price': '价格',
            'delivery': '配送',
            'shipping': '运输',
            'lost': '丢失',
            'stolen': '被盗',
            'app': '应用',
            'download': '下载',
            'install': '安装',
            'connect': '连接',
            'battery': '电池',
            'charge': '充电',
            'controller': '控制器',
            'drone': '无人机',
            'camera': '相机',
            'video': '视频',
            'photo': '照片',
            'flight': '飞行',
            'activate': '激活',
            'warranty': '保修',
            'return': '退货',
            'refund': '退款',
            
            # 德语常见短语
            'besten': '最好的',
            'service': '服务',
            'job': '工作',
            'super': '超级',
            'danke': '谢谢',
            'hallo': '你好',
            'problem': '问题',
            'hilfe': '帮助',
            'kaufen': '购买',
            'drohne': '无人机',
            'fernsteuerung': '遥控器',
            'ohne': '没有',
            'kann man': '可以',
            'wie': '如何',
            'wo': '在哪里',
            'was': '什么',
            'nicht': '不',
            'gut': '好',
            'schlecht': '坏',
            
            # 法语常见短语
            'bonjour': '你好',
            'merci': '谢谢',
            'probleme': '问题',
            'aide': '帮助',
            'acheter': '购买',
            'drone': '无人机',
            'video': '视频',
            'photo': '照片',
            'calibration': '校准',
            'mise': '设置',
            'jour': '更新',
            'bon': '好',
            'mauvais': '坏',
            'temps': '时间',
            'reponse': '回应',
            'vraiment': '真的',
            'bien': '好',
            
            # 西班牙语常见短语
            'hola': '你好',
            'gracias': '谢谢',
            'problema': '问题',
            'ayuda': '帮助',
            'comprar': '购买',
            'buenas': '你好',
            'habla': '说话',
            'español': '西班牙语',
            'cuando': '什么时候',
            'donde': '在哪里',
            
            # 意大利语
            'ciao': '你好',
            'grazie': '谢谢',
            'problema': '问题',
            'aiuto': '帮助',
            
            # 其他常见词汇
            'mini': '迷你',
            'pro': '专业版',
            'action': '运动',
            'phantom': '精灵',
            'mavic': '御',
            'osmo': '灵眸',
            'neo': '尼奥',
            'care': '关怀',
            'refresh': '更新',
            'enterprise': '企业版',
        }
        
        # 韩语翻译（基于常见模式）
        self.korean_patterns = {
            '보험': '保险',
            '입력': '输入',
            '어디': '哪里',
            '어떻게': '如何',
            '하나요': '呢',
            '하는지요': '怎么做',
            '구입': '购买',
            '활성화': '激活',
            '과정': '过程',
            '진행': '进行',
            '받아야': '需要接收',
            '코드': '代码',
            '자리': '位',
            '삼품': '商品',
            '소진': '售完',
            '주문': '订单',
            '결제': '付款',
            '배송': '配送',
            '지연': '延迟',
            '실망': '失望',
            '신속': '快速',
            '친절': '友善',
            '응대': '应对',
            '만족': '满意',
            '매우': '非常',
            '감사': '感谢',
            '합니다': '谢谢',
        }
    
    def detect_language(self, text):
        """检测文本语言"""
        if not text or pd.isna(text) or str(text).strip() == '':
            return 'unknown'
        
        text_str = str(text).strip()
        
        # 检测中文
        if re.search(r'[\u4e00-\u9fff]', text_str):
            return 'zh'
        
        # 检测韩文
        if re.search(r'[\uac00-\ud7af]', text_str):
            return 'ko'
        
        # 检测日文
        if re.search(r'[\u3040-\u309f\u30a0-\u30ff]', text_str):
            return 'ja'
        
        # 检测德语特征
        if any(word in text_str.lower() for word in ['der', 'die', 'das', 'und', 'ich', 'sie', 'ist', 'haben']):
            return 'de'
        
        # 检测法语特征
        if any(word in text_str.lower() for word in ['le', 'la', 'les', 'de', 'du', 'et', 'je', 'vous', 'est']):
            return 'fr'
        
        # 检测西班牙语特征
        if any(word in text_str.lower() for word in ['el', 'la', 'los', 'las', 'de', 'del', 'y', 'es', 'en']):
            return 'es'
        
        # 其他假设为英文
        return 'en'
    
    def translate_korean(self, text):
        """翻译韩语文本"""
        result = text
        for korean, chinese in self.korean_patterns.items():
            result = result.replace(korean, chinese)
        return result
    
    def translate_simple(self, text):
        """简单翻译功能"""
        if not text or pd.isna(text) or str(text).strip() == '':
            return ''
        
        text_str = str(text).strip()
        
        # 检测语言
        lang = self.detect_language(text_str)
        
        # 如果已经是中文，直接返回
        if lang == 'zh':
            return text_str
        
        # 韩语特殊处理
        if lang == 'ko':
            translated = self.translate_korean(text_str)
            if translated != text_str:
                return translated
        
        # 对于其他语言，进行词汇级翻译
        words = re.findall(r'\b\w+\b', text_str.lower())
        translated_parts = []
        
        # 尝试翻译每个单词
        for word in words:
            if word in self.translation_dict:
                translated_parts.append(self.translation_dict[word])
            else:
                translated_parts.append(word)
        
        # 如果有翻译结果，组合返回
        if translated_parts and any(part in self.translation_dict.values() for part in translated_parts):
            result = ' '.join(translated_parts)
            return f"{result} [原文: {text_str}]"
        
        # 如果无法翻译，返回原文
        return f"[{lang}] {text_str}"

def add_translation_to_excel():
    """为Excel文件的每个sheet添加翻译列"""
    
    input_file = '意图分类数据_按类别分组.xlsx'
    output_file = '意图分类数据_按类别分组_含翻译.xlsx'
    
    logger.info(f"开始处理文件: {input_file}")
    
    # 创建翻译器
    translator = SimpleTranslator()
    
    # 读取原Excel文件
    workbook = openpyxl.load_workbook(input_file)
    sheet_names = workbook.sheetnames
    
    logger.info(f"发现 {len(sheet_names)} 个sheet")
    
    total_translated = 0
    
    # 重新创建Excel文件，包含翻译列
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        for sheet_name in sheet_names:
            logger.info(f"处理sheet: {sheet_name}")
            
            try:
                # 读取sheet数据
                df = pd.read_excel(input_file, sheet_name=sheet_name)
                
                if 'remark' in df.columns:
                    # 添加翻译
                    translations = []
                    for idx, row in df.iterrows():
                        translation = translator.translate_simple(row['remark'])
                        translations.append(translation)
                        
                        if (idx + 1) % 50 == 0:
                            logger.info(f"  已处理 {idx + 1}/{len(df)} 条数据...")
                    
                    df['翻译'] = translations
                    total_translated += len([t for t in translations if t and t != ''])
                    
                    logger.info(f"  ✅ {sheet_name}: {len(df)}条数据处理完成")
                else:
                    logger.info(f"  ℹ️ {sheet_name}: 无remark列，仅复制数据")
                
                # 写入新文件
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 格式化
                worksheet = writer.sheets[sheet_name]
                
                # 设置表头格式
                header_font = Font(bold=True, color='FFFFFF')
                if sheet_name == '总览':
                    header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
                else:
                    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in worksheet[1]:  # 第一行是表头
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # 翻译列设置更宽的宽度
                    if len(df.columns) > 0 and column_letter == openpyxl.utils.get_column_letter(len(df.columns)):
                        adjusted_width = min(max_length + 2, 80)  # 翻译列更宽
                    else:
                        adjusted_width = min(max_length + 2, 50)
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
            except Exception as e:
                logger.error(f"处理sheet '{sheet_name}' 时出错: {e}")
                continue
    
    workbook.close()
    
    logger.info(f"✅ 翻译处理完成！")
    logger.info(f"总计处理: {total_translated} 条数据")
    logger.info(f"输出文件: {output_file}")
    
    return output_file

def main():
    """主函数"""
    try:
        logger.info("开始添加翻译列...")
        output_file = add_translation_to_excel()
        
        # 验证结果
        logger.info("验证翻译结果...")
        workbook = openpyxl.load_workbook(output_file)
        
        for sheet_name in workbook.sheetnames[:3]:  # 检查前3个sheet
            try:
                df = pd.read_excel(output_file, sheet_name=sheet_name)
                if '翻译' in df.columns and 'remark' in df.columns:
                    logger.info(f"✅ {sheet_name}: 包含翻译列 ({len(df)}行数据)")
                    # 显示样本
                    for i in range(min(3, len(df))):
                        if pd.notna(df.iloc[i]['remark']):
                            remark = str(df.iloc[i]['remark'])[:60] + '...' if len(str(df.iloc[i]['remark'])) > 60 else str(df.iloc[i]['remark'])
                            translation = str(df.iloc[i]['翻译'])[:60] + '...' if len(str(df.iloc[i]['翻译'])) > 60 else str(df.iloc[i]['翻译'])
                            logger.info(f"  原文: {remark}")
                            logger.info(f"  译文: {translation}")
                            logger.info("  ---")
                elif '翻译' in df.columns:
                    logger.info(f"✅ {sheet_name}: 包含翻译列 (无remark列)")
                else:
                    logger.warning(f"⚠️ {sheet_name}: 缺少翻译列")
            except Exception as e:
                logger.error(f"验证sheet '{sheet_name}' 时出错: {e}")
        
        workbook.close()
        
        print(f"\n🎉 翻译任务完成！")
        print(f"输出文件: {output_file}")
        print("每个包含remark列的sheet都已添加'翻译'列")
        print("翻译方式: 基于词汇字典和语言模式识别")
        
    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}")
        raise

if __name__ == "__main__":
    main()
