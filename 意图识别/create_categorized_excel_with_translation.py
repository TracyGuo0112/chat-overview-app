#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据意图分类列创建分类的Excel文件
包含翻译列的版本
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from collections import Counter
import warnings
warnings.filterwarnings('ignore')

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def analyze_intent_categories(df):
    """分析意图分类列，统计所有类别及其数量"""
    logger.info("开始分析意图分类...")
    
    # 收集所有类别
    all_categories = []
    
    for idx, row in df.iterrows():
        intent_classification = row.get('意图分类', '')
        if pd.notna(intent_classification) and str(intent_classification).strip():
            # 按逗号分割多个类别
            categories = [cat.strip() for cat in str(intent_classification).split(',') if cat.strip()]
            all_categories.extend(categories)
    
    # 统计每个类别的数量
    category_counts = Counter(all_categories)
    
    # 按数量降序排序
    sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    
    logger.info(f"发现 {len(sorted_categories)} 个不同的意图类别:")
    for i, (category, count) in enumerate(sorted_categories):
        logger.info(f"  {i}: {category} - {count}条数据")
    
    return sorted_categories

def filter_data_by_category(df, target_category):
    """筛选包含指定类别的数据"""
    filtered_rows = []
    
    for idx, row in df.iterrows():
        intent_classification = row.get('意图分类', '')
        if pd.notna(intent_classification) and str(intent_classification).strip():
            # 按逗号分割多个类别
            categories = [cat.strip() for cat in str(intent_classification).split(',') if cat.strip()]
            if target_category in categories:
                filtered_rows.append(row)
    
    if filtered_rows:
        return pd.DataFrame(filtered_rows).reset_index(drop=True)
    else:
        return pd.DataFrame()

def create_categorized_excel():
    """创建按意图分类分组的Excel文件"""
    
    input_file = '/Users/lisa/AI/意图识别/用户需求意图识别_已标注_含翻译.csv'
    output_file = '/Users/lisa/AI/意图识别/意图分类数据_按类别分组_含翻译.xlsx'
    
    logger.info(f"开始处理文件: {input_file}")
    
    try:
        # 读取CSV文件
        df = pd.read_csv(input_file)
        logger.info(f"读取到 {len(df)} 行数据")
        
        # 分析意图分类
        sorted_categories = analyze_intent_categories(df)
        
        if not sorted_categories:
            logger.error("没有发现任何意图分类数据")
            return None
        
        # 创建Excel文件
        logger.info("开始创建Excel文件...")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # 创建总览sheet
            overview_data = []
            total_records = 0
            
            for i, (category, count) in enumerate(sorted_categories):
                overview_data.append({
                    '排名': i,
                    '意图类别': category,
                    '数据量': count,
                    '占比': f"{count/len(df)*100:.1f}%"
                })
                total_records += count
            
            # 添加总计行
            overview_data.append({
                '排名': '',
                '意图类别': '总计',
                '数据量': total_records,
                '占比': f"{total_records/len(df)*100:.1f}%"
            })
            
            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name='总览', index=False)
            
            # 为每个类别创建单独的sheet
            for i, (category, count) in enumerate(sorted_categories):
                sheet_name = f"{i}_{category}"
                
                # 处理sheet名称长度限制（Excel sheet名称最多31个字符）
                if len(sheet_name) > 31:
                    sheet_name = f"{i}_{category[:25]}..."
                
                logger.info(f"创建sheet: {sheet_name} (包含 {count} 条数据)")
                
                # 筛选该类别的数据
                category_df = filter_data_by_category(df, category)
                
                if not category_df.empty:
                    # 写入数据
                    category_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"  ✅ {sheet_name}: 写入 {len(category_df)} 条数据")
                else:
                    logger.warning(f"  ⚠️ {sheet_name}: 没有找到数据")
        
        # 格式化Excel文件
        logger.info("开始格式化Excel文件...")
        format_excel_file(output_file, sorted_categories)
        
        logger.info(f"✅ Excel文件创建完成！")
        logger.info(f"输出文件: {output_file}")
        logger.info(f"包含 {len(sorted_categories) + 1} 个sheet（包括总览）")
        
        return output_file
        
    except Exception as e:
        logger.error(f"创建Excel文件时出错: {e}")
        raise

def format_excel_file(file_path, sorted_categories):
    """格式化Excel文件的样式"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        # 格式化总览sheet
        if '总览' in workbook.sheetnames:
            ws = workbook['总览']
            
            # 设置表头格式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            
            for cell in ws[1]:  # 第一行是表头
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 格式化各个类别sheet
        for i, (category, count) in enumerate(sorted_categories):
            sheet_name = f"{i}_{category}"
            if len(sheet_name) > 31:
                sheet_name = f"{i}_{category[:25]}..."
            
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                
                # 设置表头格式
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in ws[1]:  # 第一行是表头
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # 根据列的内容调整宽度
                    if column_letter in ['H', 'L']:  # remark和翻译列
                        adjusted_width = min(max_length + 2, 60)
                    else:
                        adjusted_width = min(max_length + 2, 30)
                    
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        logger.info("Excel文件格式化完成")
        
    except Exception as e:
        logger.error(f"格式化Excel文件时出错: {e}")

def main():
    """主函数"""
    try:
        logger.info("开始创建按意图分类分组的Excel文件...")
        output_file = create_categorized_excel()
        
        if output_file:
            # 验证结果
            logger.info("验证Excel文件...")
            workbook = openpyxl.load_workbook(output_file)
            sheet_names = workbook.sheetnames
            
            logger.info(f"✅ Excel文件包含 {len(sheet_names)} 个sheet:")
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(output_file, sheet_name=sheet_name)
                    logger.info(f"  - {sheet_name}: {len(df)} 行数据")
                except Exception as e:
                    logger.warning(f"  - {sheet_name}: 读取失败 - {e}")
            
            workbook.close()
            
            print(f"\n🎉 Excel文件创建完成！")
            print(f"输出文件: {output_file}")
            print("每个意图类别都有单独的sheet，按数据量降序排列")
        
    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}")
        raise

if __name__ == "__main__":
    main()

